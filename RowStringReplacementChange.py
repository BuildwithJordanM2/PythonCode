# Python Script to Update Excel Rows from "Sent" to "Python Change"
# This script allows you to load an Excel file, update specific cells in a column, and save the changes.
# It's designed to be plug-and-play and works with .xlsx files using openpyxl Python Library.

# Ensure you have the required package installed:
# pip install openpyxl, if assistance installation is needed please review https://github.com/BuildwithJordanM/python-library-installation for your specific operating system.

from openpyxl import load_workbook

# Load the Excel file
file_path = 'C:/Users/username/Downloads/blankfilename.xlsx'  # Replace with your actual file path
wb = load_workbook(file_path)

# Select the specific sheet ('TAB NAME') #Replace with your actual file tab name
sheet = wb['TAB NAME']

# Iterate through all rows in column 'T' (assuming column T is column number 20)
for row in sheet.iter_rows(min_col=20, max_col=20):  # Column T is the 20th column
    for cell in row:
        if cell.value == "Sent":
            cell.value = "Python Change"  # Update 'Sent' to 'Python Change'

# Save the modified workbook back to the same file
wb.save('C:/Users/username/Downloads/blankfilename.xlsx')  # Or you can overwrite the original file
 

Parameters:
    - pip install openpyxl Python Library.
    - file_path: str, path to the Excel file to modify. #str means string for short.
    - sheet_name: str, name of the sheet where the change will happen.
    - target_column: int, column index where the change should happen (default is 20 for column 'T'). #Column T is column 20.
    - old_value: str, the value to be replaced (default is "Sent").
    - new_value: str, the value to replace with (default is "Python Change").
