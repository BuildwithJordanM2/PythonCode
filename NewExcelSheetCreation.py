import openpyxl
from openpyxl import Workbook

# Load the workbook (Replace with a generic file path)
wb = openpyxl.load_workbook(r'C:\path\to\your\spreadsheet.xlsx')

# Get the active sheet
sh1 = wb.active

# Print the title of the active sheet
print("Active sheet title:", sh1.title)

# Print the values of the first row
for row in sh1.iter_rows(values_only=True):
    print(row)
    break  # Remove this line if you want to print all rows

# Create a new sheet (Rename as needed)
ws_new = wb.create_sheet('Sheet1')

# Save the workbook
wb.save(r'C:\path\to\your\spreadsheet.xlsx')

